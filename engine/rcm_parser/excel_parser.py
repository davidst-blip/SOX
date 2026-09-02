"""
RCM Excel parser — extracts raw rows from the Perion RCM workbook.

Returns a list of dicts (one per control row) preserving all column headers.
No LLM involved here; this is pure structural extraction.
"""

from pathlib import Path

import openpyxl


# Known header synonyms so we can find the right row even if Perion renames columns.
_HEADER_HINTS = [
    "control",
    "process",
    "description",
    "frequency",
    "owner",
    "risk",
    "assertion",
    "nature",
    "type",
    "code",
    "ref",
    "no.",
    "entity",
]

_MIN_HINT_MATCHES = 3  # a row must match at least this many hints to be the header


def _looks_like_header(row_values: list[str]) -> bool:
    lowered = [str(v).strip().lower() for v in row_values if v]
    return sum(any(h in cell for h in _HEADER_HINTS) for cell in lowered) >= _MIN_HINT_MATCHES


def parse_rcm_excel(file_path: str | Path) -> list[dict[str, str]]:
    """
    Open an RCM workbook and return all data rows as dicts.

    Strategy:
    1. Use the first sheet that has a detectable header row.
    2. Skip merged/blank rows above the header.
    3. Return {column_header: cell_value} for each non-blank data row.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        header_idx = None
        headers: list[str] = []

        for i, row in enumerate(rows):
            str_row = [str(v).strip() if v is not None else "" for v in row]
            if _looks_like_header(str_row):
                header_idx = i
                # Clean up headers: deduplicate blanks with positional fallback
                seen: dict[str, int] = {}
                for j, h in enumerate(str_row):
                    if not h:
                        h = f"col_{j}"
                    if h in seen:
                        seen[h] += 1
                        h = f"{h}_{seen[h]}"
                    else:
                        seen[h] = 0
                    headers.append(h)
                break

        if header_idx is None:
            continue  # try next sheet

        records: list[dict[str, str]] = []
        for row in rows[header_idx + 1 :]:
            str_row = [str(v).strip() if v is not None else "" for v in row]
            # Skip rows that are completely blank
            if not any(str_row):
                continue
            record = dict(zip(headers, str_row))
            records.append(record)

        if records:
            return records

    return []
