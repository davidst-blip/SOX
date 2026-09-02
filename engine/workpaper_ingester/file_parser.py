"""
Workpaper file parser — extracts text from Excel, Word, and PDF files.

Returns a list of sections: [{name, text, tables}]
No LLM involved — pure structural extraction.
"""

from pathlib import Path


def parse_workpaper(file_path: str | Path) -> list[dict]:
    """
    Parse a workpaper file into a list of sections.

    Each section:
      name  — sheet name / page number / heading
      text  — extracted text
      tables — list of 2D string arrays (rows × cols)

    Supports: .xlsx, .xlsm, .docx, .pdf
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        return _parse_excel(path)
    elif suffix == ".docx":
        return _parse_docx(path)
    elif suffix == ".pdf":
        return _parse_pdf(path)
    else:
        return [{"name": "raw", "text": f"Unsupported format: {suffix}", "tables": []}]


def _parse_excel(path: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Build a table (2D list of strings)
        table = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in rows
            if any(cell is not None for cell in row)
        ]

        # Also build plain text (non-empty cells, space-separated)
        text_lines = [" | ".join(cell for cell in row if cell) for row in table]
        text = "\n".join(text_lines)

        sections.append({"name": sheet_name, "text": text, "tables": [table]})

    return sections


def _parse_docx(path: Path) -> list[dict]:
    import docx

    doc = docx.Document(str(path))
    sections = []
    current_heading = "Document"
    current_paragraphs: list[str] = []
    current_tables: list[list[list[str]]] = []

    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

        if tag == "p":
            from docx.oxml.ns import qn
            style = element.get(qn("w:styleId"), "")
            text = "".join(t.text or "" for t in element.iter(qn("w:t")))
            if not text.strip():
                continue
            if "Heading" in style or style.startswith("h"):
                # Save previous section
                if current_paragraphs or current_tables:
                    sections.append({
                        "name": current_heading,
                        "text": "\n".join(current_paragraphs),
                        "tables": current_tables,
                    })
                current_heading = text.strip()
                current_paragraphs = []
                current_tables = []
            else:
                current_paragraphs.append(text.strip())

        elif tag == "tbl":
            from docx.oxml.ns import qn
            table_rows = []
            for tr in element.iter(qn("w:tr")):
                row = []
                for tc in tr.iter(qn("w:tc")):
                    cell_text = "".join(t.text or "" for t in tc.iter(qn("w:t")))
                    row.append(cell_text.strip())
                if any(row):
                    table_rows.append(row)
            if table_rows:
                current_tables.append(table_rows)

    # Flush last section
    if current_paragraphs or current_tables:
        sections.append({
            "name": current_heading,
            "text": "\n".join(current_paragraphs),
            "tables": current_tables,
        })

    return sections if sections else [{"name": "Document", "text": "", "tables": []}]


def _parse_pdf(path: Path) -> list[dict]:
    from pypdf import PdfReader

    reader = PdfReader(path)
    sections = []

    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        sections.append({"name": f"Page {i}", "text": text.strip(), "tables": []})

    return sections if sections else [{"name": "Page 1", "text": "", "tables": []}]
