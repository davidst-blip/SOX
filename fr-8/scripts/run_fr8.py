#!/usr/bin/env python3
"""
run_fr8.py — Generate the FR-8 Exchange Rate Working Paper.

Replicates the exact column layout and formula logic of the Perion
Daily_Exchange_Rates working paper used for SOX control FR-8.

NS-ER sheet column map (A–AK, matches original):
  A  = Date
  C/D/E   = ILS/USD NS | BOI | GAP=ABS(C-D)
  G/H/I   = ILS/EUR NS | BOI | GAP=ABS(G-H)
  K/L/M   = ILS/GBP NS | BOI | GAP=ABS(K-L)
  O/P/Q   = USD/ILS NS | 1/D (BOI-derived)   | GAP=ROUND(ABS(O-P),6)
  S/T/U   = EUR/ILS NS | 1/G (NS ILS/EUR)    | GAP=ROUND(ABS(S-T),6)
  W/X/Y   = USD/EUR NS | G/C (NS ILS rates)  | GAP=ROUND(ABS(W-X),6)
  AA/AB/AC= EUR/USD NS | C/G (NS ILS rates)  | GAP=ROUND(ABS(AA-AB),6)
  AE/AF/AG= USD/GBP NS | L/D (BOI ILS rates) | GAP=ROUND(ABS(AE-AF),6)
  AI/AJ/AK= GBP/USD NS | ROUND(D/L,6) (BOI)  | GAP=AI-AJ  (NOT abs)

Usage:
  python run_fr8.py \\
      --ns-json  /tmp/_ns_rates.json \\
      --boi-json /tmp/_boi_rates.json \\
      --period-start 2025-12-01 --period-end 2025-12-31 \\
      --output /tmp/Daily_Exchange_Rates_-_12_2025.xlsx

  python run_fr8.py --demo \\
      --period-start 2025-12-01 --period-end 2025-12-31 \\
      --output /tmp/FR8_demo.xlsx
"""

import argparse
import json
import os
import sys
from datetime import date, timedelta, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: 'openpyxl' not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BOI_URL = "https://edge.boi.gov.il/FusionDataBrowser/"
BOI_HEBREW_HEADER = "בנק ישראל - שערים יציגים"

C_DARK   = "1F3864"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_OK     = "E2EFDA"
C_FAIL   = "FFCCCC"
C_YELLOW = "FFF2CC"
C_GREY   = "F2F2F2"
C_WHITE  = "FFFFFF"


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def _date_range(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def _parse_ns_date(raw: str) -> date:
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse NS date: {raw!r}")


def _load_holidays(json_path: Path | None) -> set[date]:
    if json_path is None:
        json_path = Path(__file__).parent.parent / "references" / "israeli-holidays.json"
    if not json_path.exists():
        return set()
    with open(json_path) as f:
        data = json.load(f)
    out = set()
    for year_dates in data.get("non_publish_dates", {}).values():
        for entry in year_dates:
            try:
                out.add(date.fromisoformat(entry["date"]))
            except (KeyError, ValueError):
                pass
    return out


def _is_boi_publish_day(d: date, holidays: set[date]) -> bool:
    return d.weekday() != 5 and d not in holidays  # Saturday=5


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def load_ns_json(path: str) -> dict[tuple, float]:
    """Load NS JSON into {(date, base, quote): rate}.

    NS SuiteQL returns dates as DD/MM/YYYY; accepts YYYY-MM-DD too.
    Expects format produced by format_ns_output.py (fxsourcemethod=MANUAL,
    no self-pairs).
    """
    with open(path) as f:
        raw = json.load(f)
    records = raw if isinstance(raw, list) else raw.get("data", raw.get("records", []))
    result: dict[tuple, float] = {}
    for rec in records:
        try:
            d = _parse_ns_date(rec.get("effectivedate") or rec.get("date") or "")
            base  = (rec.get("basecurrency")        or rec.get("base_currency")        or "").upper().strip()
            quote = (rec.get("transactioncurrency") or rec.get("quote_currency")       or "").upper().strip()
            rate  = float(rec.get("exchangerate")   or rec.get("rate") or 0)
            if d and base and quote and rate and base != quote:
                result[(d, base, quote)] = rate
        except (ValueError, TypeError):
            continue
    return result


def load_boi_json(path: str) -> dict[tuple, float]:
    """Load BOI JSON into {(date, base, quote): rate}.

    fetch_boi.py stores BOI series RER_USD_ILS as base=ILS, quote=USD
    (because the rate IS ILS-per-USD, matching the NS ILS-base convention).
    """
    with open(path) as f:
        raw = json.load(f)
    observations = raw.get("observations", [])
    result: dict[tuple, float] = {}
    for obs in observations:
        try:
            d     = date.fromisoformat(obs["date"])
            base  = obs["base_currency"].upper()
            quote = obs["quote_currency"].upper()
            rate  = float(obs["rate"])
            result[(d, base, quote)] = rate
        except (KeyError, ValueError):
            continue
    return result


def _get_boi_source_mode(boi_json_path: str) -> str:
    try:
        with open(boi_json_path) as f:
            return json.load(f).get("source_mode", "unknown")
    except Exception:
        return "unknown"


# ---------------------------------------------------------------------------
# BOI carry-forward
# ---------------------------------------------------------------------------

def apply_boi_carryforward(
    boi_raw: dict[tuple, float],
    period_start: date,
    period_end: date,
) -> dict[tuple, float]:
    """For each calendar day in period, fill missing BOI rates from last published day.

    BOI does not publish on Saturdays and Israeli holidays. NS carries the
    previous day's rate. The original WP compares NS to BOI even on non-publish
    days (VLOOKUP returns N/A, but the monthly period totals stay zero). We
    replicate this by propagating the last published BOI rate forward, exactly
    as the BOI website does in its own exports.
    """
    pairs = [("ILS", "USD"), ("ILS", "EUR"), ("ILS", "GBP")]
    result: dict[tuple, float] = {}
    last: dict[tuple, float] = {}

    for d in _date_range(period_start, period_end):
        for pair in pairs:
            key = (d,) + pair
            if key in boi_raw:
                last[pair] = boi_raw[key]
                result[key] = boi_raw[key]
            elif pair in last:
                result[key] = last[pair]
    return result


# ---------------------------------------------------------------------------
# Demo data
# ---------------------------------------------------------------------------

def _generate_demo(period_start: date, period_end: date, holidays: set[date]) -> tuple[dict, dict]:
    """Return (ns_data, boi_data) with synthetic zero-gap rates.

    Rates only change on BOI publish days; weekend and holiday days carry
    the prior publish day's rate in both NS and BOI, mirroring real behaviour.
    """
    ils_usd = 3.650
    ils_eur = 3.900
    ils_gbp = 4.620
    ns: dict[tuple, float] = {}
    boi: dict[tuple, float] = {}

    publish_count = 0
    for d in _date_range(period_start, period_end):
        if _is_boi_publish_day(d, holidays):
            publish_count += 1
            u = round(ils_usd + publish_count * 0.002, 4)
            e = round(ils_eur + publish_count * 0.002, 4)
            g = round(ils_gbp + publish_count * 0.002, 4)
            boi[(d, "ILS", "USD")] = u
            boi[(d, "ILS", "EUR")] = e
            boi[(d, "ILS", "GBP")] = g
        else:
            # carry last published rates
            prev = d - timedelta(days=1)
            while prev >= period_start and (prev, "ILS", "USD") not in boi:
                prev -= timedelta(days=1)
            if (prev, "ILS", "USD") in boi:
                u = boi[(prev, "ILS", "USD")]
                e = boi[(prev, "ILS", "EUR")]
                g = boi[(prev, "ILS", "GBP")]
            else:
                u, e, g = ils_usd, ils_eur, ils_gbp

        # NS stores rates for every calendar day (same carry-forward logic as BOI)
        ns[(d, "ILS", "USD")] = u
        ns[(d, "ILS", "EUR")] = e
        ns[(d, "ILS", "GBP")] = g
        ns[(d, "USD", "ILS")] = round(1/u, 6)
        ns[(d, "EUR", "ILS")] = round(1/e, 6)
        ns[(d, "USD", "EUR")] = round(e/u, 6)
        ns[(d, "EUR", "USD")] = round(u/e, 6)
        ns[(d, "USD", "GBP")] = round(g/u, 6)
        ns[(d, "GBP", "USD")] = round(u/g, 6)

    return ns, boi


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

class Row:
    """One calendar day's worth of comparison data."""
    __slots__ = (
        "d",
        "C", "D", "E",          # ILS/USD direct
        "G", "H", "I",          # ILS/EUR direct
        "K", "L", "M",          # ILS/GBP direct
        "O", "P", "Q",          # USD/ILS cross
        "S", "T", "U",          # EUR/ILS cross
        "W", "X", "Y",          # USD/EUR cross
        "AA", "AB", "AC",       # EUR/USD cross
        "AE", "AF", "AG",       # USD/GBP cross
        "AI", "AJ", "AK",       # GBP/USD cross
        "boi_published",
    )

    def __init__(self, d: date):
        self.d = d
        for s in self.__slots__[1:]:
            setattr(self, s, None)
        self.boi_published = False


def _r(v) -> float | None:
    """Round to 6dp for cross-rate GAPs."""
    return round(v, 6) if v is not None else None


def build_rows(
    ns: dict[tuple, float],
    boi_cf: dict[tuple, float],
    boi_raw: dict[tuple, float],
    period_start: date,
    period_end: date,
) -> list[Row]:
    rows = []
    for d in _date_range(period_start, period_end):
        r = Row(d)

        # ── direct NS rates ──────────────────────────────────────────────────
        r.C = ns.get((d, "ILS", "USD"))
        r.G = ns.get((d, "ILS", "EUR"))
        r.K = ns.get((d, "ILS", "GBP"))

        # ── BOI direct rates (with carry-forward) ────────────────────────────
        r.D = boi_cf.get((d, "ILS", "USD"))
        r.H = boi_cf.get((d, "ILS", "EUR"))
        r.L = boi_cf.get((d, "ILS", "GBP"))
        r.boi_published = (d, "ILS", "USD") in boi_raw

        # ── Level 1 GAPs: NS vs BOI (direct) ────────────────────────────────
        r.E = abs(r.C - r.D) if r.C is not None and r.D is not None else None
        r.I = abs(r.G - r.H) if r.G is not None and r.H is not None else None
        r.M = abs(r.K - r.L) if r.K is not None and r.L is not None else None

        # ── NS cross-rate values ─────────────────────────────────────────────
        r.O  = ns.get((d, "USD", "ILS"))
        r.S  = ns.get((d, "EUR", "ILS"))
        r.W  = ns.get((d, "USD", "EUR"))
        r.AA = ns.get((d, "EUR", "USD"))
        r.AE = ns.get((d, "USD", "GBP"))
        r.AI = ns.get((d, "GBP", "USD"))

        # ── Level 2: cross-rate calculated values ────────────────────────────
        # P = 1/D  (BOI-derived USD/ILS)
        r.P = (1 / r.D) if r.D else None
        # T = 1/G  (NS ILS/EUR-derived)
        r.T = (1 / r.G) if r.G else None
        # X = G/C  (NS ILS/EUR ÷ NS ILS/USD)
        r.X = (r.G / r.C) if (r.G and r.C) else None
        # AB = C/G  (NS ILS/USD ÷ NS ILS/EUR)
        r.AB = (r.C / r.G) if (r.C and r.G) else None
        # AF = L/D  (BOI ILS/GBP ÷ BOI ILS/USD)
        r.AF = (r.L / r.D) if (r.L and r.D) else None
        # AJ = ROUND(D/L, 6)  (BOI ILS/USD ÷ BOI ILS/GBP)
        r.AJ = _r(r.D / r.L) if (r.D and r.L) else None

        # ── Level 2 GAPs ─────────────────────────────────────────────────────
        r.Q  = _r(abs(r.O  - r.P))  if (r.O  is not None and r.P  is not None) else None
        r.U  = _r(abs(r.S  - r.T))  if (r.S  is not None and r.T  is not None) else None
        r.Y  = _r(abs(r.W  - r.X))  if (r.W  is not None and r.X  is not None) else None
        r.AC = _r(abs(r.AA - r.AB)) if (r.AA is not None and r.AB is not None) else None
        r.AG = _r(abs(r.AE - r.AF)) if (r.AE is not None and r.AF is not None) else None
        # AK = AI - AJ  (NOT abs — direction preserved, per original)
        r.AK = (r.AI - r.AJ) if (r.AI is not None and r.AJ is not None) else None

        rows.append(r)
    return rows


# ---------------------------------------------------------------------------
# Excel styles helpers
# ---------------------------------------------------------------------------

def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic)


def _border() -> Border:
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def _auto_width(ws, mn=8, mx=18):
    for col in ws.columns:
        w = mn
        for c in col:
            if c.value is not None:
                w = max(w, min(len(str(c.value)) + 2, mx))
        ws.column_dimensions[get_column_letter(col[0].column)].width = w


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_workbook(
    rows: list[Row],
    ns: dict,
    boi_raw: dict,
    period_start: date,
    period_end: date,
    ns_json_path: str | None,
    boi_json_path: str | None,
    script_version: str,
    demo_mode: bool,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_ns_er(wb, rows, period_start, period_end, demo_mode)
    _build_boi_rates(wb, boi_raw, period_start, period_end)
    _build_ns_source_sheets(wb, ns, period_start, period_end)
    _build_ipe_evidence(wb, rows, period_start, period_end,
                        ns_json_path, boi_json_path, script_version, demo_mode)
    return wb


# ---------------------------------------------------------------------------
# NS-ER sheet — exact original layout
# ---------------------------------------------------------------------------

def _build_ns_er(wb, rows: list[Row], period_start: date, period_end: date, demo_mode: bool):
    ws = wb.create_sheet("NS - ER")
    ws.sheet_view.showGridLines = True

    period_str = f"{period_start.strftime('%B %Y')}"

    # ── Rows 1–2: sign-off block ─────────────────────────────────────────────
    ws["C1"] = BOI_URL
    ws["G1"] = "Performed by:"
    ws["G1"].font = _font(bold=True)
    ws["H1"] = ""          # name — reviewer fills in
    ws["I1"] = None        # date — reviewer fills in

    ws["G2"] = "Reviewed by:"
    ws["G2"].font = _font(bold=True)
    ws["H2"] = ""
    ws["I2"] = None

    # ── Row 3: section headers (cross-rate blocks only) ──────────────────────
    _section_header(ws, 3, "O", "Check for ILS/USD")
    _section_header(ws, 3, "S", "Check for ILS/EUR")
    _section_header(ws, 3, "W", "Check for USD/EUR")
    _section_header(ws, 3, "AA", "Check for EUR/USD")
    _section_header(ws, 3, "AE", "Check for USD/GBP")
    _section_header(ws, 3, "AI", "Check for GBP/USD")

    # ── Row 4: column headers ────────────────────────────────────────────────
    ws["A4"] = "Date"
    # Direct blocks
    for col, label in [
        ("C", "Exchange Rate Netsuite"), ("D", "Exchange Rate BOI"), ("E", "GAP"),
        ("G", "Exchange Rate Netsuite"), ("H", "Exchange Rate BOI"), ("I", "GAP"),
        ("K", "Exchange Rate Netsuite"), ("L", "Exchange Rate BOI"), ("M", "GAP"),
    ]:
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = _font(bold=True)
        ws[f"{col}4"].fill = _fill(C_LBLUE)

    # Cross-rate blocks (header matches original incl. typo)
    for col, label in [
        ("O", "Exchange Rate Netsuite"), ("P", "Exchange Rate Claculated"), ("Q", "GAP"),
        ("S", "Exchange Rate Netsuite"), ("T", "Exchange Rate Claculated"), ("U", "GAP"),
        ("W", "Exchange Rate Netsuite"), ("X", "Exchange Rate Claculated"), ("Y", "GAP"),
        ("AA", "Exchange Rate Netsuite"), ("AB", "Exchange Rate Claculated"), ("AC", "GAP"),
        ("AE", "Exchange Rate Netsuite"), ("AF", "Exchange Rate Claculated"), ("AG", "GAP"),
        ("AI", "Exchange Rate Netsuite"), ("AJ", "Exchange Rate Claculated"), ("AK", "GAP"),
    ]:
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = _font(bold=True)
        ws[f"{col}4"].fill = _fill(C_LBLUE)

    ws["A4"].font = _font(bold=True)
    ws["A4"].fill = _fill(C_LBLUE)

    # ── Row 5: currency labels + GAP totals ──────────────────────────────────
    ws["C5"] = "Curr: USD, Currency Name: US Dollars"
    ws["G5"] = "Curr: EUR, Currency Name: Euro"
    ws["K5"] = "Curr: GBP, Currency Name: British Pound"
    for lbl_col in ("C5", "G5", "K5"):
        ws[lbl_col].font = _font(italic=True, size=9)

    # Compute and write GAP sums
    gap_cols = ["E", "I", "M", "Q", "U", "Y", "AC", "AG", "AK"]
    for col in gap_cols:
        total = sum(
            getattr(r, col) for r in rows
            if getattr(r, col) is not None
        )
        ws[f"{col}5"] = round(total, 6)
        ws[f"{col}5"].font = _font(bold=True)
        ws[f"{col}5"].fill = _fill(C_OK if round(total, 6) == 0 else C_FAIL)
        ws[f"{col}5"].number_format = "0.000000"

    # ── Rows 6+: data rows ───────────────────────────────────────────────────
    for r_idx, r in enumerate(rows, 6):
        ws.cell(r_idx, 1, r.d).number_format = "DD/MM/YYYY"

        _write_rate(ws, r_idx, "C", r.C)
        _write_rate(ws, r_idx, "D", r.D)
        _write_gap(ws,  r_idx, "E", r.E)

        _write_rate(ws, r_idx, "G", r.G)
        _write_rate(ws, r_idx, "H", r.H)
        _write_gap(ws,  r_idx, "I", r.I)

        _write_rate(ws, r_idx, "K", r.K)
        _write_rate(ws, r_idx, "L", r.L)
        _write_gap(ws,  r_idx, "M", r.M)

        _write_rate(ws, r_idx, "O", r.O)
        _write_rate(ws, r_idx, "P", r.P)
        _write_gap(ws,  r_idx, "Q", r.Q, decimals=6)

        _write_rate(ws, r_idx, "S", r.S)
        _write_rate(ws, r_idx, "T", r.T)
        _write_gap(ws,  r_idx, "U", r.U, decimals=6)

        _write_rate(ws, r_idx, "W", r.W)
        _write_rate(ws, r_idx, "X", r.X)
        _write_gap(ws,  r_idx, "Y", r.Y, decimals=6)

        _write_rate(ws, r_idx, "AA", r.AA)
        _write_rate(ws, r_idx, "AB", r.AB)
        _write_gap(ws,  r_idx, "AC", r.AC, decimals=6)

        _write_rate(ws, r_idx, "AE", r.AE)
        _write_rate(ws, r_idx, "AF", r.AF)
        _write_gap(ws,  r_idx, "AG", r.AG, decimals=6)

        _write_rate(ws, r_idx, "AI", r.AI)
        _write_rate(ws, r_idx, "AJ", r.AJ)
        # AK = AI - AJ, not abs (per original formula =AI-AJ)
        _write_gap(ws,  r_idx, "AK", r.AK, decimals=6, allow_negative=True)

        # Shade non-publish days lightly
        if not r.boi_published:
            for col_letter in ("D", "H", "L"):
                ws[f"{col_letter}{r_idx}"].fill = _fill(C_GREY)

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    for col in ("B", "F", "J", "N", "R", "V", "Z", "AD", "AH"):
        ws.column_dimensions[col].width = 2
    for col in ("C","D","E","G","H","I","K","L","M",
                "O","P","Q","S","T","U","W","X","Y",
                "AA","AB","AC","AE","AF","AG","AI","AJ","AK"):
        ws.column_dimensions[col].width = 16

    ws.freeze_panes = "B6"


def _section_header(ws, row: int, col: str, text: str):
    c = ws[f"{col}{row}"]
    c.value = text
    c.font = _font(bold=True, color=C_WHITE)
    c.fill = _fill(C_BLUE)
    c.alignment = Alignment(horizontal="center")


def _write_rate(ws, row: int, col: str, value):
    if value is None:
        return
    c = ws[f"{col}{row}"]
    c.value = value
    c.number_format = "0.0000"


def _write_gap(ws, row: int, col: str, value, decimals: int = 4, allow_negative: bool = False):
    if value is None:
        return
    c = ws[f"{col}{row}"]
    c.value = value
    fmt = f"0.{'0' * decimals}"
    c.number_format = fmt
    # Highlight any non-zero gap
    threshold = 0.0001 if decimals <= 4 else 0.000001
    is_exception = abs(value) > threshold if allow_negative else (value is not None and value > threshold)
    if is_exception:
        c.fill = _fill(C_FAIL)
        c.font = _font(bold=True)


# ---------------------------------------------------------------------------
# BOI Rates sheet — matches ExchangeRates BOI format in original
# ---------------------------------------------------------------------------

def _build_boi_rates(wb, boi_raw: dict, period_start: date, period_end: date):
    ws = wb.create_sheet("ExchangeRates BOI")

    ws["A1"] = BOI_HEBREW_HEADER
    ws["A1"].font = _font(bold=True)
    ws["H1"] = BOI_URL

    period_label = (
        f"טווח תאריכים : "
        f"{period_start.strftime('%d/%m/%Y')}-{period_end.strftime('%d/%m/%Y')}"
    )
    ws["A2"] = period_label

    ws["A3"] = (
        "אם לא פורסם שער יציג ליום המבוקש, מוצג השער האחרון שפורסם לפניו"
    )
    ws["A3"].font = _font(italic=True, size=9)

    # Headers row 5 (matches original)
    for col, lbl in [("A","תאריך"), ("B","דולר ארצות הברית"),
                     ("C","ליש\"ט בריטניה"), ("D","אירו האיחוד המוניטרי האירופי")]:
        c = ws[f"{col}5"]
        c.value = lbl
        c.font = _font(bold=True)
        c.fill = _fill(C_LBLUE)

    # Data: only BOI-published dates (no carry-forward in this sheet)
    r_idx = 6
    for d in _date_range(period_start, period_end):
        usd = boi_raw.get((d, "ILS", "USD"))
        eur = boi_raw.get((d, "ILS", "EUR"))
        gbp = boi_raw.get((d, "ILS", "GBP"))
        if usd is None and eur is None and gbp is None:
            continue
        ws.cell(r_idx, 1, d).number_format = "DD/MM/YYYY"
        ws.cell(r_idx, 2, usd).number_format = "0.0000" if usd else ""
        ws.cell(r_idx, 3, gbp).number_format = "0.0000" if gbp else ""  # col C = GBP (original order)
        ws.cell(r_idx, 4, eur).number_format = "0.0000" if eur else ""  # col D = EUR
        r_idx += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# NS source sheets — one per pair direction
# ---------------------------------------------------------------------------

_NS_PAIRS = [
    ("ILS-USD",   "ILS", "USD"),
    ("ILS-EURO",  "ILS", "EUR"),
    ("ILS-GBP",   "ILS", "GBP"),
    ("USD to ILS","USD", "ILS"),
    ("EUR to ILS","EUR", "ILS"),
    ("USD to EUR","USD", "EUR"),
    ("EUR to USD","EUR", "USD"),
    ("USD to GBP","USD", "GBP"),
    ("GBP to USD","GBP", "USD"),
]


def _build_ns_source_sheets(wb, ns: dict, period_start: date, period_end: date):
    for sheet_name, base, quote in _NS_PAIRS:
        ws = wb.create_sheet(sheet_name)

        ws["A1"] = "Currency Exchange Rates"
        ws["A1"].font = _font(bold=True)
        ws["K1"] = "IL Exchange Rates Form (SS)"

        period_label = (
            f"Start Date: {period_start.strftime('%m/%d/%Y')} - "
            f"End Date: {period_end.strftime('%m/%d/%Y')}"
        )
        ws["A3"] = period_label

        for col, lbl in [("A","Base Currency"),("B","Currency"),
                         ("C","Exchange Rate"),("D","Effective Date")]:
            c = ws[f"{col}4"]
            c.value = lbl
            c.font = _font(bold=True)
            c.fill = _fill(C_LBLUE)

        r_idx = 5
        for d in _date_range(period_start, period_end):
            rate = ns.get((d, base, quote))
            if rate is None:
                continue
            ws.cell(r_idx, 1, base)
            ws.cell(r_idx, 2, quote)
            ws.cell(r_idx, 3, rate).number_format = "0.000000"
            ws.cell(r_idx, 4, d).number_format = "DD/MM/YYYY"
            r_idx += 1

        _auto_width(ws, mn=10)


# ---------------------------------------------------------------------------
# IPE Evidence sheet
# ---------------------------------------------------------------------------

def _build_ipe_evidence(wb, rows, period_start, period_end,
                        ns_path, boi_path, version, demo_mode):
    ws = wb.create_sheet("IPE Evidence")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "IPE Evidence — FR-8 Control Run Metadata"
    c.font = _font(bold=True, size=12, color=C_WHITE)
    c.fill = _fill(C_DARK)
    c.alignment = Alignment(horizontal="center")

    source = "DEMO" if demo_mode else (
        _get_boi_source_mode(boi_path) if boi_path else "N/A"
    )

    gap_cols = ["E","I","M","Q","U","Y","AC","AG","AK"]
    nonzero = [
        (r.d, col, getattr(r, col))
        for r in rows for col in gap_cols
        if getattr(r, col) is not None and abs(getattr(r, col)) > 1e-9
    ]

    meta = [
        ("Control reference", "FR-8"),
        ("Period", f"{period_start} to {period_end}"),
        ("Run timestamp (UTC)", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Script version", version),
        ("NS data source", ns_path or "DEMO"),
        ("BOI data source", boi_path or "DEMO"),
        ("BOI source mode", source),
        ("Total calendar days", (period_end - period_start).days + 1),
        ("BOI publishing days", sum(1 for r in rows if r.boi_published)),
        ("Non-zero GAPs", len(nonzero)),
        ("Overall result", "PASS" if not nonzero else f"REVIEW REQUIRED — {len(nonzero)} non-zero GAP(s)"),
        ("Formula logic — direct", "E=ABS(C-D), I=ABS(G-H), M=ABS(K-L)"),
        ("Formula logic — cross", "Q=ROUND(ABS(O-1/D),6), U=ROUND(ABS(S-1/G),6), "
                                  "Y=ROUND(ABS(W-G/C),6), AC=ROUND(ABS(AA-C/G),6), "
                                  "AG=ROUND(ABS(AE-L/D),6), AK=AI-ROUND(D/L,6)"),
        ("BOI carry-forward applied", "Yes — non-publish days use last published rate"),
    ]

    for r_idx, (lbl, val) in enumerate(meta, 4):
        lc = ws.cell(r_idx, 2, lbl)
        lc.font = _font(bold=True)
        lc.border = _border()
        vc = ws.cell(r_idx, 3, str(val))
        vc.border = _border()

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 60


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate FR-8 Exchange Rate WP")
    parser.add_argument("--ns-json",        help="Path to _ns_rates.json from NS MCP")
    parser.add_argument("--boi-json",       help="Path to _boi_rates.json from fetch_boi.py")
    parser.add_argument("--period-start",   required=True, help="YYYY-MM-DD")
    parser.add_argument("--period-end",     required=True, help="YYYY-MM-DD")
    parser.add_argument("--output",         required=True, help="Output .xlsx path")
    parser.add_argument("--holidays-json",  help="Custom Israeli holidays JSON")
    parser.add_argument("--script-version", default="v1.0.0")
    parser.add_argument("--demo",           action="store_true",
                        help="Generate with synthetic zero-gap data")
    args = parser.parse_args()

    period_start = date.fromisoformat(args.period_start)
    period_end   = date.fromisoformat(args.period_end)
    holidays     = _load_holidays(Path(args.holidays_json) if args.holidays_json else None)

    if args.demo:
        print("Running in DEMO mode — synthetic data, zero-gap scenario")
        ns, boi_raw = _generate_demo(period_start, period_end, holidays)
        ns_path = boi_path = None
    else:
        if not args.ns_json or not args.boi_json:
            print("ERROR: --ns-json and --boi-json required unless --demo", file=sys.stderr)
            sys.exit(1)
        print(f"Loading NS data from {args.ns_json}...")
        ns = load_ns_json(args.ns_json)
        print(f"  {len(ns)} NS rate records loaded")
        print(f"Loading BOI data from {args.boi_json}...")
        boi_raw = load_boi_json(args.boi_json)
        print(f"  {len(boi_raw)} BOI observations loaded")
        ns_path  = args.ns_json
        boi_path = args.boi_json

    print("Applying BOI carry-forward for non-publish days...")
    boi_cf = apply_boi_carryforward(boi_raw, period_start, period_end)

    print("Building comparison rows...")
    rows = build_rows(ns, boi_cf, boi_raw, period_start, period_end)

    # Summary
    gap_cols = ["E","I","M","Q","U","Y","AC","AG","AK"]
    nonzero = [
        (r.d, col, getattr(r, col))
        for r in rows for col in gap_cols
        if getattr(r, col) is not None and abs(getattr(r, col)) > 1e-9
    ]
    boi_days = sum(1 for r in rows if r.boi_published)
    print(f"  {len(rows)} calendar days, {boi_days} BOI publish days")
    print(f"  Non-zero GAPs: {len(nonzero)}")

    print("Building workbook...")
    wb = build_workbook(
        rows, ns, boi_raw, period_start, period_end,
        ns_path if not args.demo else None,
        boi_path if not args.demo else None,
        args.script_version,
        args.demo,
    )

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    print(f"\nOK: {args.output}")

    if nonzero:
        print(f"\nWARNING: {len(nonzero)} non-zero GAP(s) — review NS-ER sheet")
        for d, col, val in nonzero[:10]:
            print(f"  {d} col {col}: {val}")
        sys.exit(2)


if __name__ == "__main__":
    main()
